import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import load_workbook

#Loading City Data
wb = load_workbook(filename="Population.xlsx", read_only=True)
ws = wb['Worksheet']
list = []
for data in ws.iter_rows(min_row=2, min_col=1, max_row=1216, values_only=True):
    list.append(data[1])

#Autocomplete ComboBox
def search(event):
    value = event.widget.get()
    if value == '':
        city_combobox['values'] = list
    else:
        data = []
        for item in list:
            if value.lower() in item.lower():
                data.append(item)
            city_combobox['values'] = data

def entry_data():
    accepted = accept_var.get()

    if accepted == "Accepted":
        city = city_combobox.get()
        area = area_entry.get()
        landmark = landmark_entry.get()
        
        plot = plot_combobox.get()
        sqft = sqft_spinbox.get()
        availability = availability_entry.get()

        if city and area and landmark and plot and sqft and availability:
            print("City:", city, "Area:", area, "Landmark:", landmark)
            print("Plot:", plot, "Sqft:", sqft, "Availability:", availability)
            print("-----------------------------------------------------")
            messagebox.showinfo("Confirmed", "Data Entered")
        else:
            messagebox.showerror("Error", "All fields are required")
    else:
        messagebox.showerror("Error", "Please accept the terms and conditions")

#Initializing Tkinter
window = tk.Tk()
window.title('Data Entry Form')

frame = tk.Frame(window)
frame.pack()

#Saving Location Info
location_info_frame = tk.LabelFrame(frame, text="Location Information")
location_info_frame.grid(row=0, column=0, padx=20, pady=10)

city_label = tk.Label(location_info_frame, text="City")
city_combobox = ttk.Combobox(location_info_frame, values=list)
city_combobox.bind('<KeyRelease>', search)
city_label.grid(row=0, column=0)
city_combobox.grid(row=1, column=0)

area_label = tk.Label(location_info_frame, text="Area")
area_entry = tk.Entry(location_info_frame)
area_label.grid(row=0, column=2)
area_entry.grid(row=1, column=2)

landmark_label = tk.Label(location_info_frame, text="Landmark")
landmark_entry = tk.Entry(location_info_frame)
landmark_label.grid(row=0, column=3)
landmark_entry.grid(row=1, column=3)

for widget in location_info_frame.winfo_children():
    widget.grid_configure(padx=10, pady=5)

#Saving Plot Info
plot_info_frame = tk.LabelFrame(frame, text="Plot Information")
plot_info_frame.grid(row=1, column=0, sticky="news", padx=20, pady=10)

plot_label = tk.Label(plot_info_frame, text="Plot Type")
plot_combobox = ttk.Combobox(plot_info_frame, values=["Apartment", "Land", "Villa/Bungalow", "Commercial"])
plot_label.grid(row=0, column=0)
plot_combobox.grid(row=1, column=0)

sqft_label = tk.Label(plot_info_frame, text="Total Sq. Ft.")
sqft_spinbox = tk.Spinbox(plot_info_frame, from_=0, to='Infinity')
sqft_label.grid(row=0, column=1)
sqft_spinbox.grid(row=1, column=1)

availability_label = tk.Label(plot_info_frame, text="Date Of Availability")
availability_entry = tk.Entry(plot_info_frame)
availability_label.grid(row=0, column=2)
availability_entry.grid(row=1, column=2)

for widget in plot_info_frame.winfo_children():
    widget.grid_configure(padx=10, pady=5)

#Accept Terms
terms_frame = tk.LabelFrame(frame, text="Terms & Conditions")
terms_frame.grid(row=2, column=0, sticky="news", padx=20, pady=10)

accept_var = tk.StringVar(value="Not Accepted")
terms_check = tk.Checkbutton(terms_frame, text="I agree that all the information provided are true to the best of my knowledge", variable=accept_var, onvalue="Accepted", offvalue="Not Accepted")
terms_check.grid(row=0, column=0)

#Button
button = tk.Button(frame, text="Entry data", command=entry_data)
button.grid(row=3, column=0, padx=20, pady=10)

window.mainloop()